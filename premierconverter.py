"""
Automate the conversion of raw data into a specified format of data to make it more usable
"""

__version__ = '0.1.2'

#########
# Setup #
#########
# Import built-in modules
from pathlib import Path
import warnings

# Import external modules
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Configuration variables for the expected format and structure of the data
excel_extensions = ['.xlsx', '.xls', '.xlsm']

raw_struct = {
    'stem': {
        'ncols': 5,
        'chosen_cols': [0,1],
        'col_names': ['Premier_Test_Status', 'Total_Premium'],
        'col_types': [np.dtype('object'), np.dtype('float')],
    },
    'f_set': {
        'ncols': 4,
        'col_names': ['Peril_Factor', 'Relativity', 'Premium_increment', 'Premium_cumulative'],
        'col_types': [np.dtype('object')] + [np.dtype('float')] * 3,
    },
    'bp_name': 'Base Premium'
}

######################
# Workflow functions #
######################
def validate_input_filepath(input_filepath):
    """Checks on input_filepath"""    
    if not input_filepath.is_file():
        raise FileNotFoundError(
            "\n\tinput_filepath: There is no file at the input location:"
            f"\n\t'{input_filepath}'"
            "\n\tCannot read the input data"
        )
    if not input_filepath.suffix in excel_extensions:
        warnings.warn(
            f"input_filepath: The input file extension '{input_filepath.suffix}' "
            "is not a recognised Excel extension",
        )
    return(None)


def validate_output_options(out_filepath, out_sheet_name, force_overwrite):
    """
    Checks on out_filepath and out_sheet_name
    
    Returns: pd.ExcelWriter to use when saving the output
    """
    xl_writer = pd.ExcelWriter(out_filepath, engine = 'openpyxl')

    if not out_filepath.parent.is_dir():
        raise FileNotFoundError(
            f"\n\tout_filepath: The folder of the output file does not exist"
            f"Folder path: '{out_filepath.parent}'"
            "\n\tCreate the output folder before running this command"
        )

    if out_filepath.is_file():
        out_workbook = load_workbook(out_filepath)   
        if out_sheet_name in out_workbook.sheetnames and not force_overwrite:
            raise FileExistsError(
                "\n\tOutput options: Sheet already exists at the output location:"
                f"\n\tLocation: '{out_filepath}'"
                f"\n\tSheet name: '{out_sheet_name}'"
                "\n\tIf you want to overwrite it, re-run with `force_overwrite = True`"
            )
        # Set the pandas ExcelWriter to point at this workbook
        xl_writer.book = out_workbook
    else:
        if not out_filepath.suffix in excel_extensions:
            warnings.warn(
                f"out_filepath: The output file extension '{out_filepath.suffix}' "
                "is not a recognised Excel extension",
            )
    return(xl_writer)


def read_raw_data(input_filepath, input_sheet=None):
    """
    Load data from spreadsheet
    
    input_filepath: Location of the Excel file to read
    input_sheet: None or 0 for the first sheet, or the name of a sheet
    """
    # Set defaults
    if input_sheet is None:
        input_sheet = 0
    
    df_raw = pd.read_excel(
        input_filepath, sheet_name=input_sheet,
        engine="openpyxl",  # As per: https://stackoverflow.com/a/60709194
        header=None, index_col=0,
    ).rename_axis(index="Ref_num")
    
    return(df_raw)


def validate_raw_data(df_raw):
    """Checks on the loaded raw data"""
    if not (
        (df_raw.shape[1] - raw_struct['stem']['ncols']) 
        % raw_struct['f_set']['ncols'] == 0
    ):
        warnings.warn(
            f"Raw data: Incorrect number of columns: {df_raw.shape[1]}"
            "\n\tThere should be: 1 for index, "
            f"{raw_struct['stem']['ncols']} for stem section, "
            f"and by a multiple of {raw_struct['f_set']['ncols']} for factor sets"
        )
    return(None)


def get_stem_columns(df_raw):
    """Select and format the stem columns from the raw data"""
    df_stem = df_raw.iloc[
        :, raw_struct['stem']['chosen_cols']
    ].pipe(  # Rename the columns
        lambda df: df.rename(columns=dict(zip(
            df.columns, 
            raw_struct['stem']['col_names']
        )))
    )
    
    validate_stem_columns(df_stem)
    
    return(df_stem)


def validate_stem_columns(df_stem):
    """Checks on the selected stem columns"""
    if not (
        df_stem.dtypes == raw_struct['stem']['col_types']
    ).all():
        warnings.warn(
            "Stem columns: Unexpected column data types"
            f"\n\tExepcted: {raw_struct['stem']['col_types']}"
            f"\n\tActual:   {df_stem.dtypes.tolist()}"
        )
    return(None)


def get_factor_sets(df_raw):
    """Concatenate the columns in the raw data that consist of the factor sets"""
    df_fsets = pd.concat([
        # For each of the factor sets of columns
        df_raw.iloc[  # Select the columns
            :, fset_start_col:(fset_start_col + raw_struct['f_set']['ncols'])
        ].dropna(  # Remove rows that have all missing values
            how="all"
        ).pipe(lambda df: df.rename(columns=dict(zip(  # Rename columns
            df.columns, raw_struct['f_set']['col_names']
        )))).reset_index()  # Get 'Ref_num' as a column

        for fset_start_col in range(
            raw_struct['stem']['ncols'], df_raw.shape[1], raw_struct['f_set']['ncols']
        )
    ], sort=False)
    
    return(df_fsets)


def validate_factor_sets(df_fsets):
    """Checks on concatenated factor sets columns"""
    if not (
        df_fsets[raw_struct['f_set']['col_names']].dtypes == 
        raw_struct['f_set']['col_types']
    ).all():
        warnings.warn(
            "Factor sets columns: Unexpected column data types"
            f"\n\tExepcted: {raw_struct['f_set']['col_types']}"
            f"\n\tActual:   {df_fsets[raw_struct['f_set']['col_names']].dtypes.tolist()}"
        )
    return(None)


def get_implied_perils(df_fsets):
    """Get all perils in data set by looking at occurences of 'Base Premium'"""
    perils = df_fsets.Peril_Factor.drop_duplicates(  # Get only unique 'Peril_Factor' combinations
    ).to_frame().assign(  # Filter to leave only 'Base Premium' occurences
        row_to_keep=lambda df: df.Peril_Factor.str.contains(raw_struct['bp_name'])
    ).query('row_to_keep').drop(columns='row_to_keep').assign(
        # Get the 'Peril' part of 'Peril_Factor'
        Peril=lambda df: df.Peril_Factor.str.replace(raw_struct['bp_name'], "").str.strip()
    ).Peril.sort_values().to_list()
    
    return(perils)


def validate_peril_factors(df_fsets, perils_implied):
    """Checks on implied perils deduced from the factor sets"""
    if not df_fsets.Peril_Factor.str.startswith(
        tuple(perils_implied)
    ).all():
        warnings.warn(
            "Implied perils: Not every Peril_Factor starts with a Peril. "
            "Suggests the raw data format is not as expected."
        )
    
    if '' in perils_implied:
        warnings.warn(
            "Implied perils: Empty string has been implied. "
            "Suggests the raw data format is not as expected."
        )
    
    return(None)


def split_peril_factor(df_fsets, perils_implied):
    """Split the Peril_Factor column into two"""
    df_fsets_split = df_fsets.assign(
        Factor=lambda df: df.Peril_Factor.str.replace(
                '|'.join(perils_implied), ""
        ).str.strip(),
        Peril=lambda df: df.apply(
            lambda row: row.Peril_Factor.replace(row.Factor, "").strip()
            , axis=1
        )
    ).drop(columns='Peril_Factor')
    
    return(df_fsets_split)


def get_base_prems(df_fsets_split, pf_sep="_"):
    """
    Get the Base Premiums for all Ref_nums and Perils
    
    pf_sep: Seperator for Peril_Factor column names in output
    """
    df_base_prems = df_fsets_split.query(
        # Get only the Base Preimum rows
        f"Factor == '{raw_struct['bp_name']}'"
    ).assign(
        # Create Peril_Factor combination for column names
        Peril_Factor=lambda df: df.Peril + pf_sep + df.Factor,
        Custom_order=0,  # Will be used later to ensure desired column order
    ).pivot_table(
        # Pivot to 'Peril_Factor' columns and 'Ref_num' rows
        index='Ref_num',
        columns=['Peril', 'Custom_order', 'Peril_Factor'],
        values='Premium_cumulative'
    )
    
    return(df_base_prems)


def validate_base_prems(df_base_prems):
    """Checks on formatted Base Premiums"""
    if df_base_prems.isna().sum().sum() > 0:
        warnings.warn(
            "Base Premiums: Base Premium is missing for some rows and Perils. "
            "Suggests the raw data format is not as expected."
        )
    return(None)


def get_all_factor_relativities(
    df_fsets_split,
    include_factors=None,
    pf_sep='_'
):
    """
    Ensure every Ref_num has a row for every Peril, Factor combination
    Get the Relativity for all Ref_nums, Perils and Factors
    
    include_factors: If any of the factors in this list are not implied 
        in the data, then such factors are also returned in the output.
    pf_sep: Seperator for Peril_Factor column names in output
    """
    # Set defaults
    include_factors = None
    if include_factors is None:
        include_factors = []
    
    df_factors = df_fsets_split.query(
        # Get only the Factor rows
        f"Factor != '{raw_struct['bp_name']}'"
    ).drop(
        columns=['Premium_increment', 'Premium_cumulative']
    ).set_index(
        # Ensure there is one row for every combination of Ref_num, Peril, Factor
        ['Ref_num', 'Peril', 'Factor']
    ).pipe(lambda df: df.reindex(index=pd.MultiIndex.from_product([
        df.index.get_level_values('Ref_num').unique(),
        df.index.get_level_values('Peril').unique(),
        # Include additional factors if desired from the inputs
        set(df.index.get_level_values('Factor').tolist() + include_factors),
    ], names = df.index.names
    ))).sort_index().fillna({  # Any new rows need to have Relativity of 1
        'Relativity': 1.,
    }).reset_index().assign(
        # Create Peril_Factor combination for column names
        Peril_Factor=lambda df: df.Peril + pf_sep + df.Factor,
        Custom_order=1
    ).pivot_table(
        # Pivot to 'Peril_Factor' columns and 'Ref_num' rows
        index='Ref_num',
        columns=['Peril', 'Custom_order', 'Peril_Factor'],
        values='Relativity'
    )
    
    return(df_factors)


def validate_relativities(df_factors):
    """Checks on formatted Relativities"""
    if not df_factors.apply(lambda col: (col > 0)).all().all():
        warnings.warn(
            "Factor relativities: At least one relativity is below zero."
        )
    return(None)


def get_base_and_factors(df_base_prems, df_factors):
    """Combine Base Premium and Factors columns"""
    df_base_factors = df_base_prems.merge(
        df_factors,
        how='inner', left_index=True, right_index=True
    ).pipe(
        # Sort columns (uses 'Custom_order')
        lambda df: df[df.columns.sort_values()]
    )

    # Drop unwanted levels of the column MultiIndex
    df_base_factors.columns = df_base_factors.columns.get_level_values('Peril_Factor')

    return(df_base_factors)


def join_stem_to_base_factors(df_stem, df_base_factors):
    """Join formatted stem section to combined Base Premium and Factors columns"""
    df_formatted = df_stem.merge(
        df_base_factors,
        how='left', left_index=True, right_index=True
    ).fillna(0.)  # The only mising values are from 'error' rows
    
    return(df_formatted)


def save_to_workbook(df_formatted, xl_writer, out_sheet_name):
    df_formatted.to_excel(xl_writer, sheet_name=out_sheet_name)
    xl_writer.save()
    xl_writer.close()
    return(True)

######################
# Pipeline functions #
######################
def convert_df(
    df_raw,
    include_factors=None,
    pf_sep="_",
    with_validation=True,
):
    """
    Convert DataFrame of raw data into a specified format
    
    include_factors: If any of the factors in this list are not implied 
        in the data, then such factors are also returned in the output.
    pf_sep: Seperator for Peril_Factor column names in output
    with_validation: Set to False to stop optional validation checks from
        running (which might make this function run a little faster).
    """
    # Validate raw data
    if with_validation:
        validate_raw_data(df_raw)
    
    # Select and format the stem columns
    df_stem = get_stem_columns(df_raw)
    if with_validation:
        validate_stem_columns(df_stem)
    
    # Select and format the factor set columns
    df_fsets = get_factor_sets(df_raw)
    if with_validation:
        validate_factor_sets(df_fsets)
    perils_implied = get_implied_perils(df_fsets)
    if with_validation:
        validate_peril_factors(df_fsets, perils_implied)
    df_fsets_split = split_peril_factor(df_fsets, perils_implied)
    df_base_prems = get_base_prems(df_fsets_split, pf_sep)
    if with_validation:
        validate_base_prems(df_base_prems)
    df_factors = get_all_factor_relativities(
        df_fsets_split, include_factors, pf_sep
    )
    if with_validation:
        validate_relativities(df_factors)
    df_base_factors = get_base_and_factors(df_base_prems, df_factors)
    
    # Join stem and base and factor columns
    df_formatted = join_stem_to_base_factors(df_stem, df_base_factors)
    
    return(df_formatted)


def convert(
    input_filepath,
    input_sheet=None,
    out_filepath='formatted_data.xlsx',
    out_sheet_name='Sheet1',
    force_overwrite=False,
    **kwargs,
):
    """
    Load raw data from Excel, convert to specified format, and save result
    
    input_filepath: Excel file containing a sheet with the raw data
    input_sheet: None or 0 for the first sheet, or the name of a sheet
    out_filepath: Path of an Excel file to save the formatted data
    out_sheet_name: Name of the sheet to save the formatted data
    force_overwrite: Set to True if you want to overwrite the existing workbook sheet
    **kwargs: Arguments to pass to convert_df
    
    Returns: (out_filepath, out_sheet_name) if it completes
    """
    # Set defaults
    input_filepath = Path(input_filepath)
    out_filepath = Path(out_filepath)
    
    # Validate function inputs
    validate_input_filepath(input_filepath)
    xl_writer = validate_output_options(out_filepath, out_sheet_name, force_overwrite)
    
    # Load raw data
    df_raw = read_raw_data(input_filepath, input_sheet)
    
    # Get converted DataFrame
    df_formatted = convert_df(df_raw, **kwargs)
    
    # Save results to a workbook
    if save_to_workbook(df_formatted, xl_writer, out_sheet_name):
        print("Output saved")
    
    return((out_filepath, out_sheet_name))

#######################
# Reloading functions #
#######################
def load_formatted_spreadsheet(out_filepath, out_sheet_name):
    """
    Utility function to load data from output spreadsheet
    
    *Not* designed to check if there have been any changes since 
    the output sheet was created.
    """
    df_reload = pd.read_excel(
        out_filepath, sheet_name=out_sheet_name,
        engine="openpyxl",  # As per: https://stackoverflow.com/a/60709194
        index_col=[0],
    ).apply(lambda col: (
        col if col.name in raw_struct['stem']['col_names'][0]
        else col.astype('float')
    ))
    return(df_reload)
